import openpyxl

wb=openpyxl.load_workbook("7月下旬入库表.xlsx")
wb_1=openpyxl.Workbook('7月下旬入库表.xlsx')
print(wb_1.sheetnames)

# print(wb.sheetnames)
# print(wb.active)