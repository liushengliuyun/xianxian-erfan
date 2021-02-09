import os
import glob
import openpyxl

filedir =os.getcwd()


def process_merge_excel(filename):  # 对文件进行汇总
    main_file = filename[0]   #将第一个文件作为主文件
    wb = openpyxl.load_workbook(main_file)   #打开主文件的workbook对象
    sheet = wb.active   #打开活动表
    sheet.title = 'merge_result'   #将活动表名称命名为''

    for file in filename[1:]:   #将剩下文件内容添加至主文件夹活动表中
        wb2 = openpyxl.load_workbook(file)
        ac_sheet = wb2.active
        for row in ac_sheet.iter_rows(min_row=2):  #从第2行开始,按行获取所有单元格
            value = [cell.value for cell in row]    #将每行的值利用循环放置于列表中,并将其添加在主文件活动表的最后
            sheet.append(value)
    return wb


def get_file_excel(filelist):  # 获取要合并的文件
    file_list = glob.glob(os.path.join(filelist, '*.xlsx'))  #查找filelist的路径下的表格文件
    return file_list    #返回文件路径列表


if __name__ == '__main__':
    files = get_file_excel(filedir)
    wb = process_merge_excel(files)
    wb.save(r'C:\Users\Administrator\Desktop\result.xlsx')