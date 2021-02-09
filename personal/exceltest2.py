import openpyxl
wb=openpyxl.load_workbook('寒假分期计划表.xlsx')  #返回workbook对象
wb1=openpyxl.load_workbook("PyCharm.xlsx")
# print(wb1.worksheets)
# print(wb.sheetnames)   #全为workbook的属性
# print(wb.read_only)
# print(wb1.properties)

#返回worksheet对象,一遍后续对cell进行操作
ws1=wb1['快捷键的使用']
# print(wb1['快捷键的使用'])
# print(ws1.title,ws1.dimensions,ws1.max_row,ws1.min_row)
# print(ws1.rows)
# print(ws1.freeze_panes) #返回B2说明冻结了首行和首列,A2说明冻结了首行,B1说明冻结了首列
# print(ws1.values)   #返回生成器
# a=ws1.cell(row=1,column=2)
# print(a.coordinate+'='+a.value)
# for v in ws1.values:
#     print(*v)   #前面加*是什么意思

# for column in ws1.columns:
#     print(*[cell.value for cell in column ])
import os
import glob
filedir=os.getcwd()
print(filedir)
ws1.append(['肥宅快乐水','我不喜欢'])
def get_file_excel(filelist):  # 获取要合并的文件
    file_list = glob.glob(os.path.join(filelist, '*.xlsx'))  #查找filelist的路径下的表格文件
    return file_list
print(get_file_excel(filedir))
